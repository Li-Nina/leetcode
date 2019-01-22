# coding=utf-8
import xlrd
import xlwt
import chardet
import re
import jieba
import jieba.posseg as psg
import  openpyxl
xls_file = xlrd.open_workbook("project_all.xlsx")
wb = openpyxl.load_workbook("project_all.xlsx")
sheetnames = wb.sheetnames#每个工作表格的名称

meeting_list =["峰会","会议","论坛","汇报","研讨","大会","参会","汇报"]
contract_list=["合同","投标"]
develop_project_list=["项目","平台","开发"]
traffic_list=["交通费","交通"]
patent_list=["专利"]
intern_list=["实习生"]
out_list=["被退回"]
for sheetname in range(len(sheetnames)):
    table = xls_file.sheets()[sheetname]#打开第几个bu
    # people_file =xlrd.open_workbook("组名.xlsx")
    # people_table = people_file.sheets()[0]
    nrows = table.nrows

    # people_dict={}
    # p_nrows = people_table.nrows
    # for rowx in range(1,p_nrows):
    #     data = people_table.row(rowx)
    #     people_dict[data[0].value] = data[1].value
    print(nrows)#table的行数
    print(table.row(0))#table的第一行表头
    print(table.row(1))#table的第二行

    new_row = 0
    workbook = xlwt.Workbook(encoding='utf8')
    worksheet = workbook.add_sheet(sheetnames[sheetname])
    data = table.row(0)
    col_num = len(data)
    for i in range(col_num):
        worksheet.write(new_row,i,data[i].value) #单元格所在的row,column,和value.
    worksheet.write(new_row,col_num,"单据类型说明")
    new_row+=1
    name_dict={}
    #从1行开始按照行遍历
    for rowx in  range(1,nrows):
        data = table.row(rowx)
        event = data[2].value
        old_type = data[3].value
        status =  data[5].value
        print(event)  # 要分词的事件类型
        # # seg_list = psg.lcut(event)
        # # print(seg_list)  # 分词结果词seg.word,分词类型seg.flag
        # # event_word_num = len(seg_list)
        # # print(event_word_num)
        #old_type=old_type.encode('utf-8')
        if (status == "被退回" ):
            new_row += 1
            continue
        if (old_type == "差旅"):
            for i in range(col_num):
                worksheet.write(new_row,i,data[i].value) #单元格所在的row,column,和value.
            worksheet.write(new_row,col_num,old_type)
            new_row +=1
            continue
        for seg in psg.lcut(event):
            if seg.word in meeting_list:
                print(seg.word)
                new_type = "会议"
                break
            elif seg.word in contract_list:
                print(seg.word)
                new_type = "合同"
                break
            elif seg.word in traffic_list:
                print(seg.word)
                new_type = "交通费"
                break
            elif seg.word in develop_project_list:
                print(seg.word)
                new_type = "项目开发"
                break
            else:
                print(seg.word)
                new_type = old_type
        for i in range(col_num):
            worksheet.write(new_row, i, data[i].value)  # 单元格所在的row,column,和value.
        worksheet.write(new_row, col_num, new_type)
        new_row += 1
    workbook.save('./project/project_new_all.xls')