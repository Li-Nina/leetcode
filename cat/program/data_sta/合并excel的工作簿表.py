import csv
import openpyxl

wb = openpyxl.load_workbook('项目名称.xlsx')

# 获取workbook中所有的表格
sheets = wb.get_sheet_names()

print(sheets)

list1 = []
# 循环遍历所有sheet
for t in range(len(sheets)):
    sheet = wb.get_sheet_by_name(sheets[t])

    print('\n\n第' + str(t + 1) + '个sheet: ' + sheet.title + '->>>')

    # len_row代表表中有多少行数据，len_column代表excel表中有多少列数据
    len_row = sheet.max_row
    len_column = sheet.max_column
    # 合并的时候只保留第一张表的表头部分
    if t == 0:
        for i in range(1, len_row + 1):
            list2 = []
            for j in range(1, len_column + 1):
                list2.append(sheet.cell(row=i, column=j).value)
            list1.append(list2)
    else:
        for i in range(2, len_row + 1):
            list2 = []
            for j in range(1, len_column + 1):
                list2.append(sheet.cell(row=i, column=j).value)
            list1.append(list2)
# print(list1)

with open('./project/project_all.csv', 'w', newline='') as f:
    writer = csv.writer(f)
    writer.writerows(list1)
    f.close()

